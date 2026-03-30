# -*- coding: utf-8 -*-
"""
KPF Dashboard API  —  FastAPI-сервер поверх iiko REST/OLAP  (v2 optimized)
=============================================================
Запуск:
  pip install fastapi uvicorn requests pandas openpyxl
  uvicorn main:app --reload --host 0.0.0.0 --port 8000

Оптимизации v2:
  - OLAP кэш: прошлые периоды навечно, текущий день 5 мин
  - revenue + foodcost объединены в 1 OLAP-запрос
  - attendance парсится 1 раз для всех ресторанов, фильтруется из кэша
  - employees + departments кэшируются с TTL
  - /api/cache/clear — ручка для сброса кэшей
"""

import hashlib, json, logging, os, re, tempfile, time, threading, io
import xml.etree.ElementTree as ET
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime, timedelta
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests
from fastapi import FastAPI, File, HTTPException, Query, Body, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ═══════════════════════ НАСТРОЙКИ ═══════════════════════
BASE_URL = "https://starikhinkalich-co.iiko.it"
LOGIN = "nkleopa"
PASSWORD = "gazrok-3Qumfu-ceztex"
DASHBOARD_PASSWORD = "123"
TOKEN_FILE = "kpf_token.txt"
TIMEOUT = (10, 120)
DOW_NAMES = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
HK_NAMES = ["хинкали"]; HK_DOZEN = ["дюжина хинкали"]
HK_EXCL = ["упаковка", "ланч-бокс", "контейнер"]
SET_RE = re.compile(r"^(\d+)\s+(?:хинкали|хинкалей)", re.IGNORECASE)
CP_HK = "хинкали"; CP_KH = "хачапури"; GIFT_PREFIX = "подарок от шефа"
LC_CATEGORIES = [
    ("Повара",["повар","сушеф","шеф-повар","тестораскатчик","заготовщ","пельменщ"]),
    ("Админы",["администратор","управляющ","директор","менеджер"]),
    ("Официанты",["официант","хостес","бармен","кассир"]),
    ("Посудомойки",["посудомой","посудница"]),
    ("Технички",["уборщ","фея чистоты","техничес"]),
    ("Кухрабочие",["кухонный рабочий","кухраб","разнорабоч"]),
]
PLAN_FILE = "plans.json"

# ═══════════════════════ ЛОГИРОВАНИЕ ═══════════════════════
_LOG_FILE = os.path.join(os.path.dirname(__file__), "kpf_debug.log")
_fh = logging.FileHandler(_LOG_FILE, encoding="utf-8")
_fh.setFormatter(logging.Formatter("%(asctime)s  %(levelname)-8s  %(message)s"))
_ch = logging.StreamHandler()
_ch.setFormatter(logging.Formatter("%(asctime)s  %(levelname)-8s  %(message)s"))
logging.basicConfig(level=logging.INFO, handlers=[_fh, _ch])
log = logging.getLogger(__name__)

# ═══════════════════════ КЭШИ ═══════════════════════
_CACHE_TTL = 3600       # справочники: 1 час
_OLAP_CACHE_TTL = 300   # OLAP текущего дня: 5 мин
_ATT_CACHE_TTL = 300    # attendance текущего дня: 5 мин

_roles_cache: dict = {}; _roles_cache_ts: float = 0
_emp_cache: dict = {}; _emp_cache_ts: float = 0
_dept_cache: list = []; _dept_cache_ts: float = 0
_olap_cache: dict = {}; _olap_lock = threading.Lock()
_att_cache: dict = {}; _att_lock = threading.Lock()

def _is_past_range(d_from: str, d_to: str) -> bool:
    """True если весь период в прошлом (d_to < сегодня)."""
    try: return datetime.strptime(d_to, "%Y-%m-%d").date() < date.today()
    except: return False

def _olap_cache_key(payload: dict) -> str:
    return hashlib.md5(json.dumps(payload, sort_keys=True, ensure_ascii=False).encode()).hexdigest()

# ═══════════════════════ HTTP + AUTH ═══════════════════════
def _make_session():
    s = requests.Session()
    r = Retry(total=3, backoff_factor=1.5, status_forcelist=(500,502,503,504), allowed_methods=["GET","POST"])
    s.mount("https://", HTTPAdapter(max_retries=r))
    s.mount("http://", HTTPAdapter(max_retries=r))
    return s

SESSION = _make_session()

def _auth() -> Optional[str]:
    pwd = hashlib.sha1(PASSWORD.encode()).hexdigest()
    try:
        r = SESSION.post(f"{BASE_URL}/resto/api/auth", data={"login": LOGIN, "pass": pwd}, timeout=TIMEOUT, verify=False)
        if r.status_code == 200:
            t = r.text.strip()
            open(TOKEN_FILE, "w").write(t)
            log.info(f"Токен получен (len={len(t)})")
            return t
        log.error(f"Авторизация: {r.status_code} {r.text[:200]}")
    except Exception as e:
        log.error(f"Авторизация: {e}")
    return None

def get_token() -> Optional[str]:
    if os.path.exists(TOKEN_FILE):
        t = open(TOKEN_FILE).read().strip()
        if t: return t
    return _auth()

def _invalidate_token():
    if os.path.exists(TOKEN_FILE): os.remove(TOKEN_FILE)

# ═══════════════════════ OLAP С КЭШЕМ ═══════════════════════
def olap(payload: dict, tag: str = "", d_from: str = "", d_to: str = "") -> Optional[list]:
    """OLAP-запрос с кэшем. Прошлые периоды кэшируются навечно, текущий — на 5 мин."""
    ckey = _olap_cache_key(payload)
    past = _is_past_range(d_from, d_to) if d_from and d_to else False
    with _olap_lock:
        if ckey in _olap_cache:
            ts, cached = _olap_cache[ckey]
            if past or (time.time() - ts) < _OLAP_CACHE_TTL:
                log.info(f"[{tag}] CACHE HIT (past={past})")
                return cached

    token = get_token()
    if not token: return None
    url = f"{BASE_URL}/resto/api/v2/reports/olap"
    hdrs = {"Content-Type": "application/json"}
    for attempt in range(2):
        try:
            r = SESSION.post(url, params={"key": token}, json=payload, headers=hdrs, timeout=TIMEOUT, verify=False)
        except requests.RequestException as e:
            log.error(f"[{tag}] сеть: {e}"); return None
        if r.status_code == 401 and attempt == 0:
            _invalidate_token(); token = _auth()
            if not token: return None
            continue
        if r.status_code != 200:
            log.error(f"[{tag}] HTTP {r.status_code}: {r.text[:300]}"); return None
        rows = r.json().get("data", [])
        log.info(f"[{tag}] строк={len(rows)}")
        with _olap_lock: _olap_cache[ckey] = (time.time(), rows)
        return rows
    return None

def _date_filter(d_from, d_to):
    return {"filterType":"DateRange","periodType":"CUSTOM","from":d_from,"to":d_to,"includeLow":True,"includeHigh":True}

def _dept_filter(dept):
    return {"filterType":"IncludeValues","values":[dept]}

# ═══════════════════════ ЗАГРУЗКА ДАННЫХ ═══════════════════════

def fetch_revenue_and_foodcost(d_from, d_to, dept) -> tuple:
    """ОДИН запрос вместо двух: revenue + foodcost. Возвращает (rev_dict, fc_dict)."""
    rows = olap({
        "reportType": "SALES",
        "groupByRowFields": ["OpenDate.Typed"],
        "aggregateFields": ["DishDiscountSumInt", "UniqOrderId", "ProductCostBase.ProductCost"],
        "filters": {
            "OpenDate.Typed": _date_filter(d_from, d_to),
            "DeletedWithWriteoff": {"filterType": "IncludeValues", "values": ["NOT_DELETED"]},
            "Department": _dept_filter(dept),
        },
    }, "REV+FC", d_from, d_to)
    rev, fc = {}, {}
    for row in (rows or []):
        ddate = str(row.get("OpenDate.Typed", ""))[:10]
        rv = float(row.get("DishDiscountSumInt") or 0)
        orders = int(float(row.get("UniqOrderId") or 0))
        cost = float(row.get("ProductCostBase.ProductCost") or 0)
        rev[ddate] = {"rev_total": round(rv, 2), "ord_total": orders,
                      "avg_check": round(rv / orders, 2) if orders > 0 else 0.0}
        fc[ddate] = {"cost_pct": round(cost / rv * 100, 2) if rv > 0 else None}
    return rev, fc

def fetch_writeoffs(d_from, d_to, dept) -> dict:
    rows = olap({
        "reportType": "TRANSACTIONS",
        "groupByRowFields": ["DateTime.DateTyped", "Account.Name"],
        "aggregateFields": ["Sum.ResignedSum"],
        "filters": {
            "TransactionType": {"filterType": "IncludeValues", "values": ["WRITEOFF"]},
            "DateTime.Typed": _date_filter(d_from, d_to),
            "Department": _dept_filter(dept),
        },
    }, "WRITEOFF", d_from, d_to)
    result = {}
    for row in (rows or []):
        acc = str(row.get("Account.Name", "")).strip()
        if not acc.startswith("*"): continue
        ddate = str(row.get("DateTime.DateTyped", ""))[:10]
        if not ddate or ddate == "None": continue
        val = float(row.get("Sum.ResignedSum") or 0)
        if val == 0: continue
        d = result.setdefault(ddate, {"wo_total": 0.0})
        d["wo_total"] += val
    for ddate in result:
        result[ddate]["wo_total"] = round(result[ddate]["wo_total"], 2)
    return result

def fetch_cooking_time(d_from, d_to, dept) -> dict:
    rows = olap({
        "reportType": "SALES",
        "groupByRowFields": ["OpenDate.Typed", "CookingPlace"],
        "aggregateFields": ["Cooking.CookingDuration.Avg", "DishAmountInt"],
        "filters": {
            "OpenDate.Typed": _date_filter(d_from, d_to),
            "DeletedWithWriteoff": {"filterType": "IncludeValues", "values": ["NOT_DELETED"]},
            "Department": _dept_filter(dept),
        },
    }, "COOKING_TIME", d_from, d_to)
    acc = {}
    for row in (rows or []):
        ddate = str(row.get("OpenDate.Typed", ""))[:10]
        place = str(row.get("CookingPlace", "")).lower()
        try:
            secs = float(row.get("Cooking.CookingDuration.Avg") or 0)
            qty = float(row.get("DishAmountInt") or 0)
        except (TypeError, ValueError): continue
        a = acc.setdefault(ddate, {"hk_sum": 0.0, "hk_w": 0.0, "kh_sum": 0.0, "kh_w": 0.0})
        if CP_HK in place and secs >= 30 and qty > 0: a["hk_sum"] += secs * qty; a["hk_w"] += qty
        if CP_KH in place and secs >= 30 and qty > 0: a["kh_sum"] += secs * qty; a["kh_w"] += qty
    def s2t(s):
        s = int(s); return f"{s // 3600:02d}:{(s % 3600) // 60:02d}:{s % 60:02d}"
    return {
        ddate: {
            "ct_hk": s2t(a["hk_sum"] / a["hk_w"]) if a["hk_w"] > 0 else None,
            "ct_kh": s2t(a["kh_sum"] / a["kh_w"]) if a["kh_w"] > 0 else None,
        }
        for ddate, a in acc.items()
    }

def fetch_dishes(d_from, d_to, dept) -> tuple:
    """Один запрос — (khinkali_by_date, gifts_by_date)."""
    rows = olap({
        "reportType": "SALES",
        "groupByRowFields": ["OpenDate.Typed", "DishName"],
        "aggregateFields": ["DishAmountInt"],
        "filters": {
            "OpenDate.Typed": _date_filter(d_from, d_to),
            "DeletedWithWriteoff": {"filterType": "IncludeValues", "values": ["NOT_DELETED"]},
            "Department": _dept_filter(dept),
        },
    }, "DISHES", d_from, d_to)
    hk, gifts = {}, {}
    for row in (rows or []):
        ddate = str(row.get("OpenDate.Typed", ""))[:10]
        name = str(row.get("DishName", "")).strip()
        qty = int(float(row.get("DishAmountInt") or 0))
        nl = name.lower()
        if not any(k in nl for k in HK_EXCL):
            if any(k in nl for k in HK_DOZEN):
                hk.setdefault(ddate, {"hk_total": 0})["hk_total"] += qty * 12
            elif any(k in nl for k in HK_NAMES):
                m = SET_RE.match(name)
                hk.setdefault(ddate, {"hk_total": 0})["hk_total"] += qty * int(m.group(1)) if m else qty
        if nl.startswith(GIFT_PREFIX):
            gifts[ddate] = gifts.get(ddate, 0) + qty
    return hk, gifts

# ═══════════════════════ СПРАВОЧНИКИ С КЭШЕМ ═══════════════════════

def _lc_category(role_name: str) -> str:
    nl = role_name.lower()
    for cat, kws in LC_CATEGORIES:
        if any(kw in nl for kw in kws): return cat
    return "Прочие"

def _get_roles_info() -> dict:
    global _roles_cache, _roles_cache_ts
    if _roles_cache and (time.time() - _roles_cache_ts) < _CACHE_TTL: return _roles_cache
    token = get_token()
    if not token: return _roles_cache
    try: r = SESSION.get(f"{BASE_URL}/resto/api/employees/roles?key={token}", timeout=TIMEOUT, verify=False)
    except requests.RequestException as e: log.error(f"[ROLES] {e}"); return _roles_cache
    if r.status_code != 200: return _roles_cache
    roles = {}
    try:
        root = ET.fromstring(r.text)
        for role in root.findall("role"):
            rid = (role.findtext("id") or "").strip()
            name = (role.findtext("name") or "").strip()
            try: rate = float(role.findtext("paymentPerHour") or 0)
            except ValueError: rate = 0.0
            if rid: roles[rid] = {"name": name, "rate": rate}
        _roles_cache = roles; _roles_cache_ts = time.time()
        log.info(f"[ROLES] кэш обновлён: {len(roles)}")
    except ET.ParseError as e: log.error(f"[ROLES] XML: {e}")
    return _roles_cache

def _get_employee_names() -> dict:
    global _emp_cache, _emp_cache_ts
    if _emp_cache and (time.time() - _emp_cache_ts) < _CACHE_TTL: return _emp_cache
    token = get_token()
    if not token: return _emp_cache
    try: r = SESSION.get(f"{BASE_URL}/resto/api/employees?key={token}", timeout=TIMEOUT, verify=False)
    except requests.RequestException as e: log.error(f"[EMPLOYEES] {e}"); return _emp_cache
    if r.status_code != 200: return _emp_cache
    names = {}
    try:
        root = ET.fromstring(r.text)
        for emp in root.findall(".//employee"):
            eid = (emp.findtext("id") or "").strip()
            last = (emp.findtext("lastName") or "").strip()
            first = (emp.findtext("firstName") or "").strip()
            if eid: names[eid] = f"{last} {first}".strip() or eid[:12]
        _emp_cache = names; _emp_cache_ts = time.time()
        log.info(f"[EMPLOYEES] кэш обновлён: {len(names)}")
    except ET.ParseError as e: log.error(f"[EMPLOYEES] XML: {e}")
    return _emp_cache

def _parse_att_dt(s):
    s = re.sub(r"[+-]\d{2}:\d{2}$", "", s)
    return datetime.strptime(s, "%Y-%m-%dT%H:%M:%S")

# ═══════════════════════ ATTENDANCE С КЭШЕМ ═══════════════════════

def _fetch_attendance_all(d_from: str, d_to: str) -> dict:
    """Парсит attendance ОДИН раз для ВСЕХ ресторанов, кэширует.
    Возвращает: { dept_name: [list of parsed shift records] }"""
    cache_key = (d_from, d_to)
    past = _is_past_range(d_from, d_to)
    with _att_lock:
        if cache_key in _att_cache:
            ts, cached = _att_cache[cache_key]
            if past or (time.time() - ts) < _ATT_CACHE_TTL:
                log.info(f"[ATTENDANCE] CACHE HIT ({d_from}→{d_to})")
                return cached

    roles_info = _get_roles_info()
    token = get_token()
    if not token: return {}
    t0 = time.time()
    try:
        r = SESSION.get(
            f"{BASE_URL}/resto/api/employees/attendance"
            f"?from={d_from}&to={d_to}&withPaymentDetails=true&key={token}",
            timeout=TIMEOUT, verify=False)
    except requests.RequestException as e: log.error(f"[ATTENDANCE] сеть: {e}"); return {}
    if r.status_code != 200: log.error(f"[ATTENDANCE] HTTP {r.status_code}"); return {}
    try: root = ET.fromstring(r.text)
    except ET.ParseError as e: log.error(f"[ATTENDANCE] XML: {e}"); return {}

    all_depts: dict = {}
    for att in root.findall("attendance"):
        dept_name = (att.findtext("departmentName") or "").strip()
        if not dept_name: continue
        dfs = att.findtext("dateFrom") or ""
        dts = att.findtext("dateTo") or ""
        if not dfs: continue
        ddate = dfs[:10]
        role_id = (att.findtext("roleId") or "").strip()
        ri = roles_info.get(role_id, {"name": "?", "rate": 0.0})

        psum = 0.0
        pd_node = att.find("paymentDetails")
        if pd_node is not None:
            try:
                psum = (float(pd_node.findtext("regularPaymentSum") or 0)
                        + float(pd_node.findtext("overtimePayedSum") or 0)
                        + float(pd_node.findtext("otherPaymentsSum") or 0))
            except (TypeError, ValueError): pass

        hours = 0.0
        if dts:
            try: hours = (_parse_att_dt(dts) - _parse_att_dt(dfs)).total_seconds() / 3600.0
            except (ValueError, OverflowError): pass

        cost = max(psum, hours * ri["rate"] if hours > 0 else 0.0)
        no_tariff = (cost == 0 and hours > 0)

        all_depts.setdefault(dept_name, []).append({
            "ddate": ddate, "role_name": ri["name"], "category": _lc_category(ri["name"]),
            "cost": cost, "hours": hours, "no_tariff": no_tariff,
            "emp_id": (att.findtext("employeeId") or "").strip() if no_tariff else "",
        })

    log.info(f"[ATTENDANCE] parsed {sum(len(v) for v in all_depts.values())} recs, "
             f"{len(all_depts)} depts in {time.time() - t0:.1f}s")
    with _att_lock: _att_cache[cache_key] = (time.time(), all_depts)
    return all_depts

def fetch_attendance(d_from, d_to, dept) -> tuple:
    """(daily_lc, role_details, no_tariff_people) для ресторана из общего кэша."""
    records = _fetch_attendance_all(d_from, d_to).get(dept, [])
    emp_names = _get_employee_names()

    daily = {}
    role_agg = defaultdict(lambda: {
        "shifts": 0, "hours": 0.0, "cost": 0.0,
        "no_tariff_shifts": 0, "no_tariff_hours": 0.0,
    })
    no_tariff_records = []

    for rec in records:
        ddate = rec["ddate"]; rn = rec["role_name"]; cat = rec["category"]
        cost = rec["cost"]; hours = rec["hours"]; nt = rec["no_tariff"]

        if nt:
            rl = rn.lower()
            if "окл" not in rl and "цб" not in rl:
                no_tariff_records.append({
                    "employee": emp_names.get(rec["emp_id"], rec["emp_id"][:12] + "..."),
                    "role": rn, "category": cat, "date": ddate, "hours": round(hours, 1),
                })

        d = daily.setdefault(ddate, {"lc_sum": 0.0})
        d["lc_sum"] += cost
        d[f"lc_{cat}"] = d.get(f"lc_{cat}", 0.0) + cost

        ra = role_agg[rn]
        ra["shifts"] += 1; ra["hours"] += hours; ra["cost"] += cost
        if nt: ra["no_tariff_shifts"] += 1; ra["no_tariff_hours"] += hours

    role_details = [
        {"role": rn, "category": _lc_category(rn), "shifts": ra["shifts"],
         "hours": round(ra["hours"], 1), "cost": round(ra["cost"], 2),
         "no_tariff": ra["no_tariff_shifts"] > 0, "no_tariff_shifts": ra["no_tariff_shifts"],
         "no_tariff_hours": round(ra["no_tariff_hours"], 1)}
        for rn, ra in sorted(role_agg.items(), key=lambda x: -x[1]["hours"])
    ]
    return daily, role_details, no_tariff_records

def _fetch_attendance_overview(d_from, d_to) -> dict:
    """{ dept: total_lc_sum } из общего кэша attendance."""
    all_data = _fetch_attendance_all(d_from, d_to)
    return {dept: sum(r["cost"] for r in recs) for dept, recs in all_data.items()}

# ═══════════════════════ ПЛАН ═══════════════════════
def _load_plans() -> dict:
    if os.path.exists(PLAN_FILE):
        try:
            with open(PLAN_FILE, encoding="utf-8") as f: return json.load(f)
        except: pass
    return {}

def _save_plans(data: dict):
    with open(PLAN_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def _parse_excel_date_val(val) -> str:
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d") if hasattr(val, "strftime") else str(val)[:10]
    s = str(val).strip()
    m = re.match(r"(\d{2})\.(\d{2})\.(\d{4})", s)
    if m: d, mo, y = m.groups(); return f"{y}-{mo}-{d}"
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m: return s[:10]
    return ""

def fetch_departments() -> list[str]:
    global _dept_cache, _dept_cache_ts
    if _dept_cache and (time.time() - _dept_cache_ts) < _CACHE_TTL: return _dept_cache
    today = date.today()
    d_from = (today - timedelta(days=7)).strftime("%Y-%m-%d")
    d_to = (today - timedelta(days=1)).strftime("%Y-%m-%d")
    rows = olap({
        "reportType": "SALES", "groupByRowFields": ["Department"],
        "aggregateFields": ["DishDiscountSumInt"],
        "filters": {"OpenDate.Typed": _date_filter(d_from, d_to),
                    "DeletedWithWriteoff": {"filterType": "IncludeValues", "values": ["NOT_DELETED"]}},
    }, "DEPARTMENTS", d_from, d_to)
    if not rows: return _dept_cache or []
    sorted_rows = sorted(rows, key=lambda r: -float(r.get("DishDiscountSumInt", 0)))
    _dept_cache = [str(r.get("Department", "")) for r in sorted_rows if r.get("Department")]
    _dept_cache_ts = time.time()
    return _dept_cache

# ═══════════════════════ FastAPI APP ═══════════════════════
app = FastAPI(title="KPF Dashboard API", version="2.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

# ── /api/overview ──
@app.get("/api/overview")
def get_overview(date_from: str = Query(...), date_to: str = Query(...)):
    try:
        def _ovr_rev_fc():
            return olap({"reportType": "SALES", "groupByRowFields": ["Department"],
                "aggregateFields": ["DishDiscountSumInt", "UniqOrderId", "ProductCostBase.ProductCost"],
                "filters": {"OpenDate.Typed": _date_filter(date_from, date_to),
                    "DeletedWithWriteoff": {"filterType": "IncludeValues", "values": ["NOT_DELETED"]}}},
                "OVR_REV+FC", date_from, date_to) or []
        def _ovr_wo():
            return olap({"reportType": "TRANSACTIONS", "groupByRowFields": ["Department", "Account.Name"],
                "aggregateFields": ["Sum.ResignedSum"],
                "filters": {"TransactionType": {"filterType": "IncludeValues", "values": ["WRITEOFF"]},
                    "DateTime.Typed": _date_filter(date_from, date_to)}},
                "OVR_WRITEOFF", date_from, date_to) or []
        def _ovr_hk():
            return olap({"reportType": "SALES", "groupByRowFields": ["Department", "DishName"],
                "aggregateFields": ["DishAmountInt"],
                "filters": {"OpenDate.Typed": _date_filter(date_from, date_to),
                    "DeletedWithWriteoff": {"filterType": "IncludeValues", "values": ["NOT_DELETED"]}}},
                "OVR_KHINKALI", date_from, date_to) or []
        def _ovr_att():
            return _fetch_attendance_overview(date_from, date_to)

        with ThreadPoolExecutor(max_workers=4) as ex:
            f1 = ex.submit(_ovr_rev_fc); f2 = ex.submit(_ovr_wo)
            f3 = ex.submit(_ovr_hk); f4 = ex.submit(_ovr_att)
        rows_rf = f1.result(); rows_wo = f2.result(); rows_hk = f3.result(); att_by_dept = f4.result()
        plan_store = _load_plans()

        rev_map, fc_map = {}, {}
        for row in rows_rf:
            dept = str(row.get("Department", "")); rv = float(row.get("DishDiscountSumInt") or 0)
            orders = int(float(row.get("UniqOrderId") or 0)); cost = float(row.get("ProductCostBase.ProductCost") or 0)
            rev_map[dept] = {"revenue": round(rv, 2), "orders": orders, "avgCheck": round(rv / orders, 2) if orders > 0 else 0}
            fc_map[dept] = round(cost / rv * 100, 2) if rv > 0 else None

        wo_map = {}
        for row in rows_wo:
            acc = str(row.get("Account.Name", "")).strip()
            if not acc.startswith("*"): continue
            dept = str(row.get("Department", "")); val = float(row.get("Sum.ResignedSum") or 0)
            if val == 0: continue
            wo_map[dept] = round(wo_map.get(dept, 0.0) + val, 2)

        hk_map = {}
        for row in rows_hk:
            dept = str(row.get("Department", "")); name = str(row.get("DishName", "")).strip()
            qty = int(float(row.get("DishAmountInt") or 0)); nl = name.lower()
            if any(k in nl for k in HK_EXCL): continue
            if any(k in nl for k in HK_DOZEN): hk_map[dept] = hk_map.get(dept, 0) + qty * 12
            elif any(k in nl for k in HK_NAMES):
                m = SET_RE.match(name); hk_map[dept] = hk_map.get(dept, 0) + (qty * int(m.group(1)) if m else qty)

        depts = sorted(set(list(rev_map) + list(fc_map) + list(wo_map)))
        summaries = []
        for dept in depts:
            rd = rev_map.get(dept, {"revenue": 0, "orders": 0, "avgCheck": 0})
            lc = att_by_dept.get(dept, 0.0); rv = rd["revenue"]
            dp = plan_store.get(dept, {})
            sd = datetime.strptime(date_from, "%Y-%m-%d").date()
            ed = datetime.strptime(date_to, "%Y-%m-%d").date()
            pt = sum(dp.get((sd + timedelta(days=i)).strftime("%Y-%m-%d"), 0) for i in range((ed - sd).days + 1))
            summaries.append({
                "dept": dept, "revenue": rv, "orders": rd["orders"], "avgCheck": rd["avgCheck"],
                "lcSum": round(lc, 2), "lcPct": round(lc / rv * 100, 2) if rv and lc else None,
                "foodCostPct": fc_map.get(dept), "writeoffs": wo_map.get(dept, 0.0),
                "khinkali": hk_map.get(dept, 0), "plan": pt or None,
                "planPct": round(rv / pt * 100, 1) if pt and rv else None,
            })
        log.info(f"[OVERVIEW] ресторанов={len(summaries)}")
        return {"dateFrom": date_from, "dateTo": date_to, "summaries": summaries}
    except Exception as e: log.error(f"/overview: {e}"); raise HTTPException(500, str(e))

@app.get("/api/departments")
def get_departments():
    try: return {"departments": fetch_departments()}
    except Exception as e: log.error(f"/departments: {e}"); raise HTTPException(500, str(e))

@app.get("/api/debug/writeoffs")
def debug_writeoffs(dept: str = Query(...), date_from: str = Query(...), date_to: str = Query(...)):
    rows = olap({"reportType": "TRANSACTIONS", "groupByRowFields": ["DateTime.DateTyped", "Account.Name"],
        "aggregateFields": ["Sum.ResignedSum"],
        "filters": {"TransactionType": {"filterType": "IncludeValues", "values": ["WRITEOFF"]},
            "DateTime.Typed": _date_filter(date_from, date_to), "Department": _dept_filter(dept)}},
        "DEBUG_WO", date_from, date_to) or []
    accounts = {}
    for row in rows:
        acc = str(row.get("Account.Name", "")).strip()
        val = float(row.get("Sum.ResignedSum") or 0)
        accounts[acc] = round(accounts.get(acc, 0.0) + val, 2)
    return {"dept": dept, "date_from": date_from, "date_to": date_to, "total_rows": len(rows),
        "total_raw_sum": round(sum(accounts.values()), 2),
        "total_positive_sum": round(sum(v for v in accounts.values() if v > 0), 2),
        "total_starred_sum": round(sum(v for a, v in accounts.items() if a.startswith("*") and v > 0), 2),
        "accounts": dict(sorted(accounts.items(), key=lambda x: -abs(x[1])))}

@app.get("/api/daily")
def get_daily(dept: str = Query(...), date_from: str = Query(...), date_to: str = Query(...)):
    try:
        with ThreadPoolExecutor(max_workers=4) as ex:
            f1 = ex.submit(fetch_revenue_and_foodcost, date_from, date_to, dept)
            f2 = ex.submit(fetch_writeoffs, date_from, date_to, dept)
            f3 = ex.submit(fetch_cooking_time, date_from, date_to, dept)
            f4 = ex.submit(fetch_dishes, date_from, date_to, dept)
            f5 = ex.submit(fetch_attendance, date_from, date_to, dept)
        rev, fc = f1.result(); wo = f2.result(); ct = f3.result()
        hk, gifts = f4.result(); att_daily, _, _ = f5.result()
    except Exception as e: log.error(f"/daily: {e}"); raise HTTPException(500, str(e))

    plan_store = _load_plans()
    start = datetime.strptime(date_from, "%Y-%m-%d").date()
    end = datetime.strptime(date_to, "%Y-%m-%d").date()
    days = []; cur = start
    while cur <= end:
        ddate = cur.strftime("%Y-%m-%d"); dow = DOW_NAMES[cur.weekday()]
        r = rev.get(ddate, {}); w = wo.get(ddate, {}); f = fc.get(ddate, {})
        c = ct.get(ddate, {}); h = hk.get(ddate, {}); g = gifts.get(ddate, 0)
        a = att_daily.get(ddate, {})
        rv = r.get("rev_total", 0.0); lc = a.get("lc_sum", 0.0)
        pv = plan_store.get(dept, {}).get(ddate)
        days.append({
            "date": ddate, "dow": dow, "plan": pv,
            "planPct": round(rv / pv * 100, 1) if pv and rv else None,
            "revenue": rv, "orders": r.get("ord_total", 0), "avgCheck": r.get("avg_check", 0.0),
            "foodCostPct": f.get("cost_pct"),
            "lcSum": round(lc, 2), "lcPct": round(lc / rv * 100, 2) if rv and lc else None,
            "writeoffs": w.get("wo_total", 0.0), "khinkali": h.get("hk_total", 0), "gifts": g,
            "ctHk": c.get("ct_hk"), "ctKh": c.get("ct_kh"),
            "lcCooks": round(a.get("lc_Повара", 0.0), 2),
            "lcAdmins": round(a.get("lc_Админы", 0.0), 2),
            "lcWaiters": round(a.get("lc_Официанты", 0.0), 2),
            "lcDishwashers": round(a.get("lc_Посудомойки", 0.0), 2),
            "lcCleaners": round(a.get("lc_Технички", 0.0), 2),
            "lcKitchen": round(a.get("lc_Кухрабочие", 0.0), 2),
            "lcOther": round(a.get("lc_Прочие", 0.0), 2),
        })
        cur += timedelta(days=1)
    return {"dept": dept, "dateFrom": date_from, "dateTo": date_to, "days": days}

@app.get("/api/labor")
def get_labor(dept: str = Query(...), date_from: str = Query(...), date_to: str = Query(...)):
    try: _, role_details, no_tariff = fetch_attendance(date_from, date_to, dept)
    except Exception as e: raise HTTPException(500, str(e))
    return {"dept": dept, "roles": role_details, "noTariff": no_tariff}

@app.get("/api/plan")
def get_plan(dept: str = Query(...), date_from: str = Query(...), date_to: str = Query(...)):
    store = _load_plans(); dp = store.get(dept, {})
    start = datetime.strptime(date_from, "%Y-%m-%d").date()
    end = datetime.strptime(date_to, "%Y-%m-%d").date()
    result = {}; cur = start
    while cur <= end: d = cur.strftime("%Y-%m-%d"); result[d] = dp.get(d); cur += timedelta(days=1)
    return {"dept": dept, "plan": result}

@app.get("/api/plan/template")
def download_plan_template(date_from: str = Query(...), date_to: str = Query(...)):
    depts = fetch_departments()
    if not depts: raise HTTPException(500, "Не удалось получить список ресторанов")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "План"
    HF = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HFont = Font(color="FFFFFF", bold=True, size=10, name="Arial")
    PF = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    CA = Alignment(horizontal="center", vertical="center", wrap_text=True)
    BD = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    c = ws.cell(row=1, column=1, value="Дата"); c.font = HFont; c.fill = HF; c.alignment = CA; c.border = BD
    ws.column_dimensions["A"].width = 22
    for i, dept in enumerate(depts, 2):
        c = ws.cell(row=1, column=i, value=dept); c.font = HFont; c.fill = HF; c.alignment = CA; c.border = BD
        ws.column_dimensions[get_column_letter(i)].width = max(16, len(dept) + 2)
    ws.row_dimensions[1].height = 32
    start = datetime.strptime(date_from, "%Y-%m-%d").date()
    end = datetime.strptime(date_to, "%Y-%m-%d").date()
    ri = 1; cur = start
    while cur <= end:
        ri += 1; dow = DOW_NAMES[cur.weekday()]; label = f"{cur.strftime('%d.%m.%Y')}, {dow}"
        ca = ws.cell(row=ri, column=1, value=label); ca.alignment = CA; ca.border = BD; ca.font = Font(size=10, name="Arial")
        for col in range(2, len(depts) + 2):
            cc = ws.cell(row=ri, column=col, value=None); cc.fill = PF; cc.alignment = CA; cc.border = BD; cc.number_format = '#,##0'
        ws.row_dimensions[ri].height = 18; cur += timedelta(days=1)
    ws.freeze_panes = "A2"
    ws2 = wb.create_sheet("Инструкция")
    for r, t in enumerate(["Формат файла", "", "Лист «План»:",
        "  Столбец A — Дата (формат ДД.ММ.ГГГГ, ДД.ММ.ГГГГ, Пн и т.д.)",
        "  Столбцы B, C, ... — Рестораны (заголовки = точные названия из iiko)",
        "  Ячейки плана — числа без пробелов и знаков валюты", "", "Пример:",
        "  01.04.2026, Ср | 180000 | 220000 | ...", "",
        "Пустые ячейки = план не задан (старое значение не изменится).",
        "Загрузка перезаписывает только заполненные ячейки."], 1):
        ws2.cell(row=r, column=1, value=t).font = Font(size=11)
    ws2.column_dimensions["A"].width = 70
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="plan_template_{date_from}_{date_to}.xlsx"'})

@app.post("/api/plan/upload")
async def upload_plan(file: UploadFile = File(...)):
    contents = await file.read(); tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(contents); tmp_path = tmp.name
        wb = openpyxl.load_workbook(tmp_path, data_only=True); store = _load_plans(); saved = 0; preview = []
        for sn in wb.sheetnames:
            ws = wb[sn]; rows = list(ws.iter_rows(values_only=True))
            if not rows: continue
            headers = [str(c or "").strip() for c in rows[0]]; fh = headers[0].lower()
            if len(headers) > 2 and fh in ("дата", "date", ""):
                for row in rows[1:]:
                    dv = row[0]
                    if not dv: continue
                    ds = _parse_excel_date_val(dv)
                    if not ds: continue
                    for i, dn in enumerate(headers[1:]):
                        if not dn: continue
                        raw = row[i + 1] if i + 1 < len(row) else None
                        if raw is None or raw == "": continue
                        try:
                            v = float(str(raw).replace("\xa0", "").replace(" ", "").replace(",", ".").replace("₽", "").replace("р.", ""))
                            store.setdefault(dn, {})[ds] = v; saved += 1
                            if len(preview) < 100: preview.append({"date": ds, "dept": dn, "plan": v})
                        except ValueError: pass
            else:
                dn = sn.strip(); sr = 1 if fh in ("дата", "date", "план", "plan") else 0
                for row in rows[sr:]:
                    dv = row[0] if len(row) > 0 else None; pr = row[1] if len(row) > 1 else None
                    if not dv or pr is None: continue
                    ds = _parse_excel_date_val(dv)
                    if not ds: continue
                    try:
                        v = float(str(pr).replace("\xa0", "").replace(" ", "").replace(",", ".").replace("₽", "").replace("р.", ""))
                        store.setdefault(dn, {})[ds] = v; saved += 1
                        if len(preview) < 100: preview.append({"date": ds, "dept": dn, "plan": v})
                    except ValueError: pass
        _save_plans(store); log.info(f"[PLAN UPLOAD] сохранено {saved} из {file.filename}")
        return {"ok": True, "saved": saved, "preview": preview}
    except Exception as e: log.error(f"[PLAN UPLOAD] {e}"); raise HTTPException(400, str(e))
    finally:
        if tmp_path:
            try: os.unlink(tmp_path)
            except: pass

@app.post("/api/plan")
def set_plan(body: dict = Body(...)):
    dept = body.get("dept"); plan = body.get("plan", {})
    if not dept or not plan: raise HTTPException(400, "dept и plan обязательны")
    store = _load_plans()
    if dept not in store: store[dept] = {}
    for d, v in plan.items():
        if v is None: store[dept].pop(d, None)
        else: store[dept][d] = float(v)
    _save_plans(store); log.info(f"[PLAN] {len(plan)} дней для {dept}")
    return {"ok": True, "saved": len(plan)}

@app.get("/api/auth/check")
def auth_check():
    if get_token(): return {"ok": True}
    raise HTTPException(401, "Не удалось авторизоваться в iiko")

@app.post("/api/auth/login")
def auth_login(body: dict):
    if body.get("login") == LOGIN and body.get("password") == DASHBOARD_PASSWORD:
        token = _auth()
        if token: return {"ok": True, "name": "Управляющий"}
        raise HTTPException(502, "iiko недоступен")
    raise HTTPException(401, "Неверный логин или пароль")

@app.get("/api/debug/cooking_places")
def debug_cooking_places(dept: str = Query(...), date_from: str = Query(...), date_to: str = Query(...)):
    rows = olap({"reportType": "SALES", "groupByRowFields": ["CookingPlace"], "aggregateFields": ["DishAmountInt"],
        "filters": {"OpenDate.Typed": _date_filter(date_from, date_to),
            "DeletedWithWriteoff": {"filterType": "IncludeValues", "values": ["NOT_DELETED"]},
            "Department": _dept_filter(dept)}}, "DEBUG_CP", date_from, date_to)
    places = sorted(set(str(r.get("CookingPlace", "")).strip() for r in (rows or [])))
    return {"dept": dept, "all_places": places,
        "matched_hk": [p for p in places if CP_HK in p.lower()],
        "matched_kh": [p for p in places if CP_KH in p.lower()],
        "cp_hk_pattern": CP_HK, "cp_kh_pattern": CP_KH}

@app.post("/api/cache/clear")
def clear_cache():
    global _olap_cache, _att_cache, _roles_cache, _roles_cache_ts
    global _emp_cache, _emp_cache_ts, _dept_cache, _dept_cache_ts
    with _olap_lock: co = len(_olap_cache); _olap_cache.clear()
    with _att_lock: ca = len(_att_cache); _att_cache.clear()
    _roles_cache = {}; _roles_cache_ts = 0; _emp_cache = {}; _emp_cache_ts = 0
    _dept_cache = []; _dept_cache_ts = 0
    log.info(f"[CACHE CLEAR] olap={co}, att={ca}")
    return {"ok": True, "cleared": {"olap": co, "attendance": ca}}